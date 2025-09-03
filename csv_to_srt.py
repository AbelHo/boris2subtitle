import csv
from datetime import timedelta
import openpyxl
import os

def seconds_to_srt_time(seconds):
    td = timedelta(seconds=float(seconds))
    total_seconds = int(td.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    secs = total_seconds % 60
    millis = int((td.total_seconds() - total_seconds) * 1000)
    return f"{hours:02}:{minutes:02}:{secs:02},{millis:03}"

def csv_to_srt(input_file, output_srt):
    ext = os.path.splitext(input_file)[1].lower()
    subtitles = []
    if ext == '.csv':
        with open(input_file, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for idx, row in enumerate(reader, 1):
                start = seconds_to_srt_time(row["Start (s)"])
                stop = seconds_to_srt_time(row["Stop (s)"])
                text = f"{row['Behavior type'].strip().upper()}: {row['Behavior'].strip()}"
                subtitles.append(f"{idx}\n{start} --> {stop}\n{text}\n")
    elif ext == '.xlsx':
        wb = openpyxl.load_workbook(input_file, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        start_idx = headers.index('Start (s)')
        stop_idx = headers.index('Stop (s)')
        type_idx = headers.index('Behavior type')
        beh_idx = headers.index('Behavior')
        for idx, row in enumerate(ws.iter_rows(min_row=2), 1):
            start = seconds_to_srt_time(row[start_idx].value)
            stop = seconds_to_srt_time(row[stop_idx].value)
            text = f"{str(row[type_idx].value).strip().upper()}: {str(row[beh_idx].value).strip()}"
            subtitles.append(f"{idx}\n{start} --> {stop}\n{text}\n")
    else:
        raise ValueError('Unsupported file type. Only .csv and .xlsx are accepted.')
    with open(output_srt, "w", encoding="utf-8") as srtfile:
        srtfile.write("\n".join(subtitles))
    print(f"SRT file created: {output_srt}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) == 3:
        csv_to_srt(sys.argv[1], sys.argv[2])
    else:
        # Default behavior for original files
        csv_to_srt("AA.csv", "AA.srt")
